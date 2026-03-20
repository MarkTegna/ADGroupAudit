"""Report generator for AD Group Audit.

Generates an Excel spreadsheet of protected groups, their membership, and dates.
Reads from the database only - no AD queries.
"""

import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ad_group_audit.db_service import DatabaseService

logger = logging.getLogger("ad_group_audit")

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ACTIVE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
REMOVED_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")


def generate_report(db: DatabaseService, output_path: str = None) -> str:
    """Generate an Excel report of protected groups and membership.

    Args:
        db: Connected DatabaseService instance.
        output_path: Optional output file path. Auto-generated if None.

    Returns:
        Path to the generated Excel file.
    """
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d-%H-%M")
        output_path = f"ad-group-audit-report-{timestamp}.xlsx"

    wb = Workbook()
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _build_summary_sheet(ws_summary, db)

    # Detail sheet - all membership records
    ws_detail = wb.create_sheet("Membership Detail")
    _build_detail_sheet(ws_detail, db)

    wb.save(output_path)
    logger.info("Report saved to: %s", output_path)
    print(f"Report saved to: {output_path}")
    return output_path


def _build_summary_sheet(ws, db: DatabaseService):
    """Build the summary sheet with protected group overview."""
    headers = ["Group Name", "Domain", "DN", "Active Members", "Removed Members", "Total Records"]
    _write_header_row(ws, headers)

    cursor = db.conn.cursor()
    cursor.execute("""
        SELECT g.name, g.domain, g.dn,
            (SELECT COUNT(*) FROM Membership m WHERE m.group_dn = g.dn AND m.first_not_seen IS NULL) as active_count,
            (SELECT COUNT(*) FROM Membership m WHERE m.group_dn = g.dn AND m.first_not_seen IS NOT NULL) as removed_count,
            (SELECT COUNT(*) FROM Membership m WHERE m.group_dn = g.dn) as total_count
        FROM Groups g
        WHERE g.is_protected = 1
        ORDER BY g.domain, g.name
    """)

    row = 2
    for rec in cursor.fetchall():
        ws.cell(row=row, column=1, value=rec.name)
        ws.cell(row=row, column=2, value=rec.domain)
        ws.cell(row=row, column=3, value=rec.dn)
        ws.cell(row=row, column=4, value=rec.active_count)
        ws.cell(row=row, column=5, value=rec.removed_count)
        ws.cell(row=row, column=6, value=rec.total_count)
        row += 1

    _auto_width(ws, headers)


def _build_detail_sheet(ws, db: DatabaseService):
    """Build the detail sheet with all membership records for protected groups."""
    headers = ["Group Name", "Domain", "Member DN", "First Seen", "First Not Seen", "Status"]
    _write_header_row(ws, headers)

    cursor = db.conn.cursor()
    cursor.execute("""
        SELECT g.name, g.domain, m.member_dn, m.first_seen, m.first_not_seen
        FROM Membership m
        INNER JOIN Groups g ON m.group_dn = g.dn AND m.domain = g.domain
        WHERE g.is_protected = 1
        ORDER BY g.domain, g.name, CASE WHEN m.first_not_seen IS NULL THEN 0 ELSE 1 END, m.member_dn
    """)

    row = 2
    for rec in cursor.fetchall():
        is_active = rec.first_not_seen is None
        status = "Active" if is_active else "Removed"
        fill = ACTIVE_FILL if is_active else REMOVED_FILL

        ws.cell(row=row, column=1, value=rec.name)
        ws.cell(row=row, column=2, value=rec.domain)
        ws.cell(row=row, column=3, value=rec.member_dn)
        ws.cell(row=row, column=4, value=str(rec.first_seen) if rec.first_seen else "")
        ws.cell(row=row, column=5, value=str(rec.first_not_seen) if rec.first_not_seen else "")
        ws.cell(row=row, column=6, value=status)

        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = fill

        row += 1

    _auto_width(ws, headers)


def _write_header_row(ws, headers):
    """Write a formatted header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _auto_width(ws, headers):
    """Auto-size columns based on header width (minimum)."""
    for col, header in enumerate(headers, 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = max(len(header) + 4, 15)
